# test
### Daiki Tobita  2022/05/27 ###

import numpy as np
import openpyxl
from scipy import signal
from scipy import fftpack
from matplotlib import pyplot as plt
from openpyxl.chart import ScatterChart, Reference, Series
from tqdm import tqdm, trange


### ファイル読み込み
readfile = input("Input readfile name: ")
wb = openpyxl.load_workbook(readfile + ".xlsx")
ws = wb.worksheets[0]

### 定数宣言
q = ws.cell(2, 8).value  # 流量
n = ws.cell(2, 9).value  # 粗度係数
g = 9.81  # 重力加速度
h2diff = 0.01

# ループ１
for datanum in range(6, 99999):

    # セルが空白なら抜ける
    if ws.cell(datanum + 1, 2).value is None:
        break

    i = ws.cell(datanum, 6).value  # 勾配
    width1 = ws.cell(datanum - 1, 4).value  # 河川幅1
    width2 = ws.cell(datanum, 4).value
    h1 = ws.cell(datanum - 1, 8).value  # 水深1
    a1 = width1 * h1  # 断面積1
    r1 = a1 / (2 * h1 + width1)  # 径深1
    hc = (q**2 / (g * width1**2)) ** (1 / 3)  # 限界水深
    d1 = ws.cell(datanum - 1, 2).value
    d2 = ws.cell(datanum, 2).value
    v1 = q / a1
    l = (d2 - d1) * 1000
    const1 = q**2 / (2 * g)
    const2 = n**2 * q**2 / 2

    h2 = h1
    sign = 1
    num = 0

    if datanum == 6:
        fr1 = v1 * (g * h1) ** (-0.5)
        if fr1 > 1:
            ws.cell(datanum - 1, 9, "斜流")
        elif fr1 < 1:
            ws.cell(datanum - 1, 9, "常流")
        else:
            ws.cell(datanum - 1, 9, "限界水深")

    a2 = width2 * h2
    r2 = a2 / (2 * h2 + width2)

    lhs = h2 + (i * l)
    rhs = (
        h1
        + (const1 * (a1 ** (-2) - a2 ** (-2)))
        + (const2 * ((r1 ** (-4 / 3) * a1 ** (-2)) + (r2 ** (-4 / 3) * a2 ** (-2)))) * l
    )
    flg = abs(lhs - rhs)

    # ループ2
    while True:

        h2 += sign * h2diff

        a2 = width2 * h2
        r2 = a2 / (2 * h2 + width2)

        lhs = h2 + (i * l)
        rhs = (
            h1
            + (const1 * (a1 ** (-2) - a2 ** (-2)))
            + (const2 * ((r1 ** (-4 / 3) * a1 ** (-2)) + (r2 ** (-4 / 3) * a2 ** (-2))))
            * l
        )

        # 2回目且つ水深減
        if num == 1 and sign < 0:
            sign = -1

        # 最初　水深変化決定
        if num == 0:
            if abs(lhs - rhs) < flg:
                sign = 1
            else:
                sign = -1
                h2 = h1 + sign * h2diff
                a2 = width2 * h2
                r2 = a2 / (2 * h2 + width2)
                lhs = h2 + (i * l)
                rhs = (
                    h1
                    + (const1 * (a1 ** (-2) - a2 ** (-2)))
                    + (
                        const2
                        * (
                            (r1 ** (-4 / 3) * a1 ** (-2))
                            + (r2 ** (-4 / 3) * a2 ** (-2))
                        )
                    )
                    * l
                )

        # ひとつ先のh2と比較して、良くなるなら続行、悪くなるならループを抜ける
        if abs(lhs - rhs) < flg:
            flg = abs(lhs - rhs)
        else:

            h2 -= sign * h2diff
            break

        num += 1

    ws.cell(datanum, 8, h2)
    v2 = q / a2

    fr2 = v2 * (g * h2) ** (-0.5)
    if fr2 > 1:
        ws.cell(datanum, 9, "斜流")
    elif fr2 < 1:
        ws.cell(datanum, 9, "常流")
    else:
        ws.cell(datanum, 9, "限界水深")

    # 限界水深
    ws.cell(datanum, 10, hc)

    wb.save(readfile + ".xlsx")
